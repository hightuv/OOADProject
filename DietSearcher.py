import openpyxl


class DietSearcher:
    def __init__(self):
        self.search_key = ''
        self.workbook = 0
        self.num_row = 24
        self.num_col = 8
        self.error = 0
        self.error_message = ''
        self.response = ''

    def search(self, year, month, menu):
        try:
            self.workbook = openpyxl.load_workbook('./Result/%d년 %d월 식단표.xlsx' % (year, month))
        except ValueError:
            self.error += 1
            self.error_message += '식단표 로드하는 것을 실패했습니다.'
        row = 1
        col = 65
        response = []
        for ws in self.workbook.worksheets:
            for i in range(self.num_col):
                for j in range(self.num_row):
                    cell = ws[chr(col + i) + str(row + j)].value
                    if cell == menu:
                        response.append(ws[chr(col + i) + str(5)].value)
        for r in response:
            self.response += r + ' '

    def get_error_message(self):
        return self.error_message

    def get_error(self):
        return self.error

    def get_response(self):
        return self.response
