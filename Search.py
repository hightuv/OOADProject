import openpyxl


class Search:
    def __init__(self):
        self.search_key = ''
        self.workbook = 0
        self.num_row = 24
        self.num_col = 8
        self.result = ''

    def do_search(self, year, month, menu):
        self.workbook = openpyxl.load_workbook('./Result/%d년 %d월 식단표.xlsx' % (year, month))
        row = 1
        col = 65
        result = []
        for ws in self.workbook.worksheets:
            for i in range(self.num_col):
                for j in range(self.num_row):
                    cell = ws[chr(col + i) + str(row + j)].value
                    if cell == menu:
                        result.append(ws[chr(col + i) + str(5)].value)
        for r in result:
            self.result += r + ' '

    def get_result(self):
        return self.result
